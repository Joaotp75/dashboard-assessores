import os
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
import plotly.express as px

# Título do Dashboard
st.title("Dashboard de Relatórios de Comissões - Consolidado")

# Upload de arquivos pela barra lateral
st.sidebar.header("Upload de Arquivos")
uploaded_files = st.sidebar.file_uploader(
    "Selecione arquivos Excel (.xlsx)",
    type=['xlsx'],
    accept_multiple_files=True
)

# Processamento dos arquivos uploadados
if uploaded_files:
    dados_consolidados = []
    for arquivo in uploaded_files:
        wb = load_workbook(arquivo, data_only=True)
        codigo_assessor = arquivo.name[:6]  # Extrai o código do nome do arquivo

        for aba in wb.sheetnames:
            ws = wb[aba]

            for i in range(2, ws.max_row + 1):
                data = ws.cell(row=i, column=1).value
                produto = ws.cell(row=i, column=2).value
                valor_movimentacao = ws.cell(row=i, column=3).value
                roa = ws.cell(row=i, column=4).value
                comissao_bruta = ws.cell(row=i, column=5).value

                if any([data, produto, valor_movimentacao, roa, comissao_bruta]):
                    dados_consolidados.append({
                        'Código': codigo_assessor,
                        'Mês': aba,
                        'Data': data,
                        'Produto': produto,
                        'Valor Movimentação': valor_movimentacao,
                        'ROA': roa,
                        'Comissão Bruta': comissao_bruta
                    })

    df_consolidado = pd.DataFrame(dados_consolidados)
else:
    st.warning("Nenhum arquivo foi enviado.")
    df_consolidado = pd.DataFrame()

# Exibe filtros e resumo
if not df_consolidado.empty:
    st.sidebar.header("Filtros")

    codigos_unicos = ['Todos'] + list(df_consolidado['Código'].unique())
    codigo_selecionado = st.sidebar.selectbox("Código do Assessor", codigos_unicos)

    meses_unicos = ['Todos'] + list(df_consolidado['Mês'].unique())
    mes_selecionado = st.sidebar.selectbox("Mês do Lançamento", meses_unicos)

    # Filtragem dos dados
    if codigo_selecionado == 'Todos':
        df_filtrado = df_consolidado
    else:
        df_filtrado = df_consolidado[df_consolidado['Código'] == codigo_selecionado]

    if mes_selecionado != 'Todos':
        df_filtrado = df_filtrado[df_filtrado['Mês'] == mes_selecionado]

    st.write(f"Lançamentos para {codigo_selecionado} - {mes_selecionado}")

    if not df_filtrado.empty:
        # Tabela Resumo
        captacao_total = df_filtrado['Valor Movimentação'].sum()
        comissao_total = df_filtrado['Comissão Bruta'].sum()
        roa_medio = comissao_total / captacao_total if captacao_total > 0 else 0

        resumo = pd.DataFrame({
            'Captação Total': [captacao_total],
            'Comissão Total': [comissao_total],
            'ROA Médio (%)': [roa_medio * 100]
        })

        st.subheader("Resumo do Período")
        st.dataframe(resumo.style.format({
            'Captação Total': "R$ {:,.2f}",
            'Comissão Total': "R$ {:,.2f}",
            'ROA Médio (%)': "{:.2f}%"
        }))

        # Análise Temporal - Tabela Mensal
        df_temporal = df_filtrado.groupby('Mês').agg({
            'Valor Movimentação': 'sum',
            'Comissão Bruta': 'sum'
        }).reset_index()

        df_temporal['ROA Médio'] = (
            df_temporal['Comissão Bruta'] / df_temporal['Valor Movimentação']
        ) * 100

        # Ordenação Correta dos Meses
        ordem_meses = [
            'Janeiro', 'Fevereiro', 'Março', 'Abril',
            'Maio', 'Junho', 'Julho', 'Agosto',
            'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]
        df_temporal['Mês'] = pd.Categorical(
            df_temporal['Mês'],
            categories=ordem_meses,
            ordered=True
        )
        df_temporal = df_temporal.sort_values('Mês')

        st.subheader("Análise Temporal por Mês")
        st.dataframe(df_temporal.style.format({
            'Valor Movimentação': "R$ {:,.2f}",
            'Comissão Bruta': "R$ {:,.2f}",
            'ROA Médio': "{:.2f}%"
        }))

        # Gráfico Temporal - Movimentação
        fig_mov = px.line(
            df_temporal,
            x='Mês',
            y='Valor Movimentação',
            markers=True,
            title="Evolução da Movimentação por Mês"
        )
        fig_mov.update_layout(
            xaxis_title="Mês",
            yaxis_title="Valor Movimentação (R$)"
        )
        st.plotly_chart(fig_mov)

        # Gráfico Temporal - Comissão
        fig_com = px.line(
            df_temporal,
            x='Mês',
            y='Comissão Bruta',
            markers=True,
            title="Evolução da Comissão por Mês"
        )
        fig_com.update_layout(
            xaxis_title="Mês",
            yaxis_title="Comissão Bruta (R$)"
        )
        st.plotly_chart(fig_com)

        # Exibe Tabela de Operações
        st.subheader("Operações no Período")
        st.dataframe(df_filtrado.style.format({
            'Valor Movimentação': "R$ {:,.2f}",
            'Comissão Bruta': "R$ {:,.2f}",
            'ROA': "{:.2%}"
        }))

        # Número de operações
        st.write(f"Número de Operações: {len(df_filtrado)}")
    else:
        st.warning("Nenhuma operação encontrada para este filtro.")
